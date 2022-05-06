import openpyxl
import os
import platform
import zipfile
import string
import win32com.client
import glob
import time


o = win32com.client.Dispatch("Excel.Application")
o.Visible = False
letter = string.ascii_uppercase
my_system = platform.uname()
format_to_open =[".xlsx", ".xlsm", ".xltx", "xltm","XLSX"]
format_to_transfrom = [".xls", ".xlsb", ".xltx", ".xltm",".xlt",".xml",".xlam",".xla",".xlw",".xlr"]


def type_of_file(folder_path):
    folder_path = str(folder_path)
    find_first = folder_path.find(":")
    find_lest = folder_path.rfind(":")
    if find_first > 0 and find_lest != find_first:
        folder_path= folder_path[0:find_lest-1]
        print(folder_path)
    elif find_lest == find_first:
        print(folder_path)

class Folder_And_File:
    def __init__(self, folder_pth=None, file_path=None):
        self.folder_path = str(folder_pth).strip()
        self.getSystem = str(my_system.system)
        self.file_path = str(file_path).strip()

        if len(self.folder_path) > 0 and len(self.file_path) == 0:
            Folder_And_File.check_folder(self)

        elif len(self.folder_path) == 0 and len(self.file_path) == 0:
            raise Exception("Sorry you didn't write the location of the file or the folder")

        elif len(self.folder_path) == 0 and len(self.file_path) > 0:
            pass

        if file_path.find(":") > 0 and len(folder_pth) > 0:
            Folder_And_File.check_folder(self)

    def check_folder(self):
        if self.getSystem == "Windows":
            folder_pth = self.folder_path.replace('\\',"/")
        find_first = self.folder_path.find(":")
        find_lest = self.folder_path.rfind(":")
        if find_first > 0 and find_lest != find_first:
            self.folder_path = self.folder_path[0:find_lest - 1]
        user_action(self.folder_path)


class Action_On_File(Folder_And_File):
    def __init__(self, search_for_data, folder, file):
        self.open_file_action = openpyxl.load_workbook(folder + "/" + file)
        self.active_file_to_action = self.open_file_action.active
        self.value_search = search_for_data
        self.get_all_sheets_name = self.open_file_action.sheetnames

    def search_for_value(self):
        for sheet_look in self.get_all_sheets_name:
            self.active_sheet_search = self.open_file_action[sheet_look]
            self.get_max_row = self.active_sheet_search.max_row
            self.get_max_col = self.active_sheet_search.max_column
            self.flag = 0

            for a in range(0,len(letter)):
                for row in range(1, self.get_max_row):
                    self.value_look_for = self.active_sheet_search["{}{}".format(letter[a], row)].value
                    if self.value_look_for == self.value_search:
                        self.letter_return = letter[a]
                        self.row_found = row
                        self.return_sheet = sheet_look
                        self.flag = 1
                        return self.letter_return, self.row_found, self.return_sheet

            if self.flag == 0:
                for a in range(0, len(letter)):
                    for b in range(0, len(letter)):
                            for row in range(1, self.get_max_row):
                                    self.value_look_for = self.active_sheet_search["{}{}".format(letter[a] + letter[b], row)].value
                                    if self.value_look_for == self.value_search:
                                        self.letter_return = letter[a] + letter[b]
                                        self.row_found = row
                                        self.return_sheet = sheet_look
                                        self.flag = 1
                                        return self.letter_return, self.row_found, self.return_sheet
            elif self.flag == 0:
                for a in range(0, len(letter)):
                    for b in range(0, len(letter)):
                        for c in range(0, len(letter)):
                            for row in range(1, self.get_max_row):
                                    if a < 26 and b < 26 and c < 26 and self.flag ==0:
                                        self.value_look_for = self.active_sheet_search["{}{}".format(letter[a]+ letter[b]+ letter[c], row)].value
                                        if self.value_look_for == self.value_search:
                                            self.letter_return = letter[a] + letter[b]
                                            self.row_found = row
                                            self.flag = 1
                                            self.return_sheet = sheet_look
                                            return self.letter_return, self.row_found, self.return_sheet
        self.active_sheet_search.close()

def user_action(folder_pth): # user select hes action
    ask_for_action = True
    while ask_for_action != "q":
        ask_for_action = input("choose the action you want to do \n S(search)\n T(change Type from xls to xlsx)\n C(copy data that you search to empty workbook)\n Q(quit)\nwhat action do you want to do ?")
        ask_for_action = ask_for_action.strip()
        if ask_for_action != "q" and len(ask_for_action) > 0:
            ask_for_action = ask_for_action.strip().lower()
            action(ask_for_action,folder_pth)

def action(ask_for_action,folder_pth):
    switcher = {
        "s": search,
        "t": change_type_file_to_xlsx,
        "c": copy_data,
    }
    get_func_action = switcher.get(ask_for_action, "choose another action")
    if switcher.get(ask_for_action) != None:
        result = get_func_action(folder_pth, file=None)
    elif switcher.get(ask_for_action)== None:
        print("you write {}".format(ask_for_action), "but the option are can be only s, t, c ", get_func_action)


def search(folder_pth, file):
    search_for_data = input("what are you looking for ? ")
    count_file_folder = 0
    file_not_in_format = 0
    if len(os.listdir(folder_pth)) == 0:
        raise Exception("you don't have files in the folder {}".format(folder_pth))
    for file_in_folder in os.listdir(folder_pth):
        count_file_folder += 1
        file_in_folder_end = os.path.splitext(file_in_folder)[-1].lower()
        if file_in_folder_end in format_to_open:
            print(file_in_folder)
            try:
                letter_found = Action_On_File(search_for_data,folder_pth, file_in_folder)
                letter_found.search_for_value()
                which_sheet = letter_found.return_sheet
                which_col = letter_found.letter_return
                which_row = letter_found.row_found
                open_files = openpyxl.load_workbook(folder_pth + "/" + file_in_folder)
                what_kind_of_action = True
                ask_user_for_action = True
                while ask_user_for_action != "Yes" or ask_user_for_action !="No":
                    ask_user_for_action = input("what action do you want to do on the found data yes or no ?").capitalize()
                    print(ask_user_for_action)
                    if ask_user_for_action == "Yes":
                        while what_kind_of_action != "None":
                            what_kind_of_action = input("for copy row to new workbook(cr) for delete row(dr) for get information on the row(ir) or None").capitalize()
                            print(what_kind_of_action)
                            if what_kind_of_action == "YES":
                                action_after_find_data(which_col, which_row, which_sheet)
                            elif what_kind_of_action == "None":
                                print("your data its on ",which_col, which_row, which_sheet)
                    elif ask_user_for_action == "No":
                        print("your data its on ", which_col, which_row, which_sheet)

            except:
                print("the value is not found in file ", file_in_folder)

        elif file_in_folder in format_to_transfrom:
            file_not_in_format +=1
            user_choose_the_action = 0
            while (ask_user_for_action != "NO" or ask_user_for_action != "YES") and user_choose_the_action != 1:
                ask_chanbge_format = input("do you want to change to format of files\n to change write yes \n if you dont want to change write No ?").strip().upper()
                if ask_chanbge_format == "YES":
                    change_type_file_to_xlsx(folder_pth, file=file_in_folder)
                    user_choose_the_action = 1
                elif ask_chanbge_format == "NO":
                    user_choose_the_action = 1
                    pass

    if len(os.listdir(folder_pth)) == file_not_in_format:
            raise Exception("your file in the folder are not xlsx")


def change_type_file_to_xlsx(folder_pth=None, file=None):
    if file == None and folder_pth != None:
        if my_system.system == "Windows":
            input_dir = str(folder_pth).replace("/","\\")
        for format_file in format_to_transfrom:
            print(input_dir + "/*{}".format(format_file))
            files_xls = glob.glob(input_dir + "/*{}".format(format_file))
            output_dir = input_dir
            for filename in files_xls:
                file_change = os.path.basename(filename)
                output = output_dir + '/' + file_change.replace('.xls', '.xlsx')
                wb = o.Workbooks.Open(filename)
                wb.ActiveSheet.SaveAs(output, 51)
                wb.Close(True)
                time.sleep(5)

    if file != None and folder_pth == None:
            pass



    for format_file_to_check in format_to_transfrom:
            files_xls = glob.glob(input_dir + "/*{}".format(format_file_to_check))
            for files_in_folder in files_xls:
                try:
                    filename_xlsx = files_in_folder.replace("{}".format(format_file_to_check), ".xlsx")
                    print(files_in_folder)
                    print(filename_xlsx)
                    if os.path.exists(files_in_folder) and os.path.exists(filename_xlsx):  # if both of them are TRUE so delete the old file
                        os.remove(files_in_folder)
                except:
                    print("problem in ", files_in_folder)




def get_name_and_folder(name_file, folder):
        if name_file in folder:
            file_to_open = folder
            return file_to_open
        else:
            file_to_open = folder + "/" + name_file
            return file_to_open


def copy_data(folder_pth):
    print("next")


def open_excel(file_to_open):
            open_excel_file = openpyxl.load_workbook(file_to_open)
            active_file = open_excel_file.active
            MaxRow = active_file.max_row
            print(MaxRow)


def action_after_find_data(user_action,which_col, which_row, which_sheet):
    action_to_do={

    }




opens = Folder_And_File(r"C:\Users\Kobi Malul\Desktop\נכס בודד ר\CHANGE_Name","")
